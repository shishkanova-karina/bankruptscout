"""Загрузка списка ИНН из Excel-файлов и строковых переменных."""

from __future__ import annotations

import os
import re

from openpyxl import load_workbook


def clean_inn(raw: str) -> str:
    return re.sub(r"\D", "", raw)


def _unique(seq: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for item in seq:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def read_inns_from_xlsx(
    filepath: str,
    sheet_name: str | None = None,
    column: str | None = None,
) -> list[str]:
    """Считывает ИНН из xlsx; берёт ячейки, в которых >= 10 цифр."""
    if not filepath or not os.path.isfile(filepath):
        raise FileNotFoundError(filepath)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    inns: list[str] = []
    col_idx = ord(column.upper()) - ord("A") if column and column.isalpha() else None

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        idx = col_idx if col_idx is not None else 0
        if len(row) <= idx or row[idx] is None:
            continue
        val = clean_inn(str(row[idx]))
        if len(val) >= 10:
            inns.append(val)

    wb.close()
    return _unique(inns)


def read_inns_from_string(raw: str | None) -> list[str]:
    if not raw:
        return []
    chunks = re.split(r"[,\s;]+", raw.strip())
    return _unique([clean_inn(c) for c in chunks if len(clean_inn(c)) >= 10])
